import os
import mysql.connector
import openpyxl
import datetime


class SQL:
    def __init__(self,username,password,file=""):
        self.username = username
        self.password = password
        self.file = file
    def login(self,drop_list = None):
        mydb = mysql.connector.connect(host='localhost',user=self.username ,passwd=self.password)
        mycursor = mydb.cursor()
        if drop_list is not None:
            mycursor.execute("SHOW DATABASES")
            drop_list.config(completevalues = [m for i in mycursor.fetchall() for m in i])
            drop_list.current(0)
                
    def read(self,txt_fnc,datab_name,file_name):
        try:
            os.remove("Cache")
        except Exception:
            pass
        try:    
            mydb = mysql.connector.connect(host='localhost',user=self.username ,passwd=self.password ,database = datab_name)
        except Exception:
            mydb = mysql.connector.connect(host='localhost',user=self.username ,passwd=self.password)
            mycursor = mydb.cursor()
            mycursor.execute(f"CREATE DATABASE IF NOT EXISTS {datab_name}")
            mycursor.execute(f"USE {datab_name}")

        mycursor = mydb.cursor()
        wb = openpyxl.load_workbook(self.file)
        t1 = wb.active
        row = t1.max_row
        column = t1.max_column
        num_rows = 0

        for iter_row in range(1,row+1):
            list_row = []
            for iter_col in range(1,column+1):
                list_row.append(t1.cell(iter_row,iter_col).value)
            if any(list_row):
                if type(t1.cell(iter_row,iter_col)).__name__ == "MergedCell":
                    del list_row
                    continue
                if (num_rows == 0) and (list_row.count(None) == len(list_row)-1):
                    del list_row
                    continue

                with open("Cache","a") as file:
                    file.write(f"{list_row}\n".replace("None","'NULL'"))
                    num_rows += 1
                    num_col = len(list_row)
                  
                  
        len_col = []
        type_list = []
        for val in range(num_col):
            cnt_1 = True
            with open("Cache","r") as cache_file:
                for list_val in cache_file.readlines()[1:]:
                    list_val1 = eval(list_val)
                    if list_val1[val] == "NULL":
                        continue
                    if cnt_1:
                        type_list.append(type(list_val1[val]))
                        len_col.append(len(str(list_val1[val])))
                        cnt_1 = False
                        continue
                    if (type_list[val] != type(list_val1[val])):
                        if  float in [type_list[val],type(list_val1[val])]:
                            type_list[val] = float
                        else:    
                            type_list[val] = str    
                    if len_col[val] < len(str(list_val1[val])):
                        len_col[val] = len(str(list_val1[val]))    
        for m in type_list:
            if m == int:
                if int(len_col[type_list.index(m)]) < 10:
                    type_list[type_list.index(m)] = f"INT({len_col[type_list.index(m)]})"
                else:
                    type_list[type_list.index(m)] = f"BIGINT({len_col[type_list.index(m)]})"
            elif m == float:
                type_list[type_list.index(m)] = "FLOAT"           
            elif "datetime" in str(m):
                type_list[type_list.index(m)] = "DATETIME"
            else:
                type_list[type_list.index(m)] = f"VARCHAR({len_col[type_list.index(m)]})"
            
        cnt_2  = True
        with open("Cache","r") as cache_file:
            for list_item in cache_file.readlines():
                list_item = eval(list_item)            
                if  cnt_2: 
                    command = ""
                    for j in range(len(list_item)):
                        list_item[j] = list_item[j].replace(" ","").replace(".","").replace("(","").replace(")","")
                        try:
                            command += list_item[j] + " " + type_list[j] + ","
                        except Exception:
                            type_list.append("VARCHAR(10)")
                            command += list_item[j] + " " + type_list[j] + ","    
                    cr_exe = f"CREATE TABLE {file_name}({command.rstrip(',')})"
                    txt_fnc(cr_exe)
                    print(cr_exe)
                    mycursor.execute(cr_exe)
                    cnt_2 = False
                else:
                    for q in range(len(list_item)):
                        if list_item[q] != "NULL":
                            list_item[q] = f'"{list_item[q]}"' 
                        else:
                            pass    
                    exe = f"INSERT INTO {file_name} VALUES({','.join(list_item)})"
                    txt_fnc(exe)
                    print(exe) 
                    mycursor.execute(exe)
                    mydb.commit()
        os.remove("Cache")            
        mydb.close()
        txt_fnc(f'Successfully created table "{file_name}"')
        print("Successfully created!")

               


